import openpyxl

# Load the Excel workbook
workbook = openpyxl.load_workbook("Meanwhile Reserving Model_v1.16 - Broken Links.xlsx", data_only=False)

# Iterate through sheets
all_sheets = ""
for sheet_name in workbook.sheetnames:
	sheet = workbook[sheet_name]
	all_sheets += sheet_name + "\n"

	print(f"Excel Tab, Sheet Name:", sheet_name)

	formulas = ""

	# Iterate through cells
	#for row in sheet.iter_cols():
	if sheet_name == "Calculation Engine": # This will get only one tab from the excel. If you want more files then comment this if statement
		for row in sheet.iter_rows():
			for cell in row:
				if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
					sFormula = f"Cell {cell.coordinate}: {cell.value}, "
					formulas += sFormula + "\t"
				#elif cell.value and isinstance(cell.value, str) and cell.value !="None": 
				elif cell.value and cell.value !="None": 
					sFormula = f"Cell {cell.coordinate}: {cell.value}, "
					formulas += sFormula + "\t"
				elif cell.value =="None" or cell.value == "": 
					formulas += "\n\n"
			
			formulas += "\n\n"

		# Write formulas to a file named after the sheet
		file_name = f"{sheet_name}_row.txt"
		with open(file_name, "w") as file:
			file.write(f"Excel Tab, Sheet Name: {sheet_name}" + "\n")
			file.write(formulas)
			#print(f"Formulas written to {file_name}")

		print("Tables present in the file are = ", sheet.tables)
		for table in sheet.tables.values():
			print("Tables present in the sheet = ", table)
			
# Write formulas to a file named after the sheet
file_name = f"all_sheets.txt"
with open(file_name, "w") as file:
	file.write("Total number of sheets in the workbook are : " + "\n")
	file.write(all_sheets)

print(f"Formulas written to {file_name}")